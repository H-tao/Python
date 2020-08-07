#!/usr/bin python
# -*- coding: utf-8 -*-
# __time__ = 2020/7/20


"""
脚本工作思路:
# 0. 输入云龙流水线用户名和密码
# 1. 取得流水线的所有信息（服务名、要爬取的检查数据项、数据项对应EXCEL表格的Sheet、行、列）
# 2. 爬取检查数据项
# 3. 存储数据项到所需要的表格
    存储一条流水线的信息需要:
    a). 爬取的检查数据项值
    b). 检查数据项对应到表格的表单、行、列数据
"""

from settings import GO_CHECK_DATA, GO_EXCEL_INDEX, C_CHECK_DATA, C_EXCEL_INDEX
from settings import Python_CHECK_DATA, Python_EXCEL_INDEX, Only_Python_CHECK_DATA
from settings import FINAL_OUTPUT_EXCEL_PATH, PIPELINES_OUTPUT_EXCEL_PATH, EXECUTABLE_PATH, INPUT_EXCEL_PATH
from settings import CHECK_SET, SHOW_WINDOWS
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from selenium.webdriver.common.by import By
import selenium.webdriver.support.expected_conditions as EC
import selenium.webdriver.support.ui as ui
from selenium import webdriver
import pandas as pd
import numpy as np
import openpyxl
import time
import re
import os
import requests
import json
from tempfile import TemporaryFile

USERNAME = "swx944363"
PASSWORD = ''


class CloudPiplineWebUI:
    def __init__(self):
        self.browser = None
        self.current_url = ''
        self.llt_projects = []
        self.sheets_dict = {}
        self.wb = openpyxl.load_workbook(INPUT_EXCEL_PATH)
        self.init_browser()
        self.init_integration_excel_df()

    def init_browser(self):
        chrome_options = webdriver.ChromeOptions()
        if not SHOW_WINDOWS:
            chrome_options.add_argument('--headless')
            chrome_options.add_argument('--disable-gpu')  # 上面三行代码就是为了将Chrome不弹出界面，实现无界面爬取
        else:
            chrome_options = None
        self.browser = webdriver.Chrome(
            # 驱动安装路径
            executable_path=EXECUTABLE_PATH,
            options=chrome_options,
        )

    def init_integration_excel_df(self):
        sheets = ["Go服务", "Python服务", "C&C++服务"]
        for s in sheets:
            info_integration_df = pd.read_excel(INPUT_EXCEL_PATH, sheet_name=s)
            self.sheets_dict[s] = info_integration_df

    def open_url(self, url):
        try:
            self.browser.get(url)
            print("Open CloudAnalytics url %s success!" % url)
        except Exception as e:
            print(e)

    def login(self):
        print("Login ...")
        driver = self.browser
        self.is_visible('//*[@id="uid"]')
        element = driver.find_element_by_xpath('//*[@id="uid"]')
        element.send_keys(USERNAME)
        element = driver.find_element_by_xpath('//*[@id="password"]')
        element.send_keys(PASSWORD)
        element = driver.find_element_by_xpath('//*[@name="Submit"]')
        self.element_click(element)

    def select_all_pipelines(self):
        time.sleep(4)  # 等待10秒，待元素加载完全
        self.click_element('//*[@id="filterByService"]/div/div/input')  # 点击- 查询策略框
        self.click_element('//*[@id="filterByService"]/div/ul/ul/li[4]')  # 点击选择- 所有的流水线
        time.sleep(2)

    # 一直等待某个元素出现，默认超时10秒
    def is_visible(self, locator, timeout=10):
        try:
            ui.WebDriverWait(self.browser, timeout).until(EC.visibility_of_element_located((By.XPATH, locator)))
            return True
        except TimeoutException:
            return False

    # 一直等待某个元素消失，默认超时10秒
    def is_not_visible(self, locator, timeout=10):
        try:
            ui.WebDriverWait(self.browser, timeout).until_not(EC.visibility_of_element_located((By.XPATH, locator)))
            return True
        except TimeoutException:
            return False

    # 一直等待某元素可点击，默认超时10秒
    def is_clickable(self, locator, timeout=10):
        try:
            ui.WebDriverWait(self.browser, timeout).until(EC.element_to_be_clickable((By.XPATH, locator)))
            return True
        except TimeoutException:
            return False

    @staticmethod
    def element_click(element):
        try:
            element.click()
        except Exception as e:  # 输出报错原因
            print(element.id, e)

    def click_element(self, xpath):
        self.is_visible(xpath)  # 等待元素出现
        element = self.browser.find_element_by_xpath(xpath)  # 定位元素
        self.is_clickable(xpath)  # 等待元素可点击
        self.element_click(element)  # 点击元素

    def send_keys(self, xpath, keys):
        self.is_visible(xpath)
        element = self.browser.find_element_by_xpath(xpath)
        element.send_keys(keys)

    def switch_to_main(self):
        driver = self.browser
        driver.close()  # 关闭当前窗口
        driver.switch_to.window(driver.window_handles[0])  # 切换主窗口

    def quit(self):
        self.browser.quit()

    def search_one(self, pipeline_name):
        driver = self.browser
        element = driver.find_element_by_xpath('//*[@id="filterByName"]/div/input')  # 搜索输入框
        element.clear()  # 清空搜索框原有数据
        element.send_keys(pipeline_name)  # 输入流水线名
        fail_count = 3  # 失败重试次数
        while fail_count:
            try:
                self.click_element('//*[@id="filterByName"]/div/span[1]')  # 点击搜索按钮
                pipeline_xpath = f'//*[@id="goDetail_{pipeline_name}"]'  # 搜索之后，流水线xpath会根据流水线名动态生成
                self.click_element(pipeline_xpath)  # 点击流水线
                driver.switch_to.window(driver.window_handles[-1])  # 切换新窗口
                self.click_element('//*[@id="detailQualityID_Build"]')  # 点击- 质量检测报告按钮
                self.is_visible('//*[@id="quality-dialog"]/div/div/stage-all-quality/div/div/div')  # 等待质量检测报告加载
            except NoSuchElementException:
                print("打开新流水线窗口错误，重试中......")
                driver.switch_to.window(driver.window_handles[0])  # 切换回主窗口
                fail_count -= 1
                time.sleep(2)
                continue
            else:
                self.current_url = driver.current_url
                break

    def get_report(self):
        driver = self.browser
        # 抓取所有数据
        while True:
            try:
                self.is_visible('//*[@id="quality-dialog"]/div/div/stage-all-quality/div/div/ul/li/ul')
                element = driver.find_element_by_xpath(
                    '//*[@id="quality-dialog"]/div/div/stage-all-quality/div/div/ul/li/ul')
                result = element.text.split('\n')
                print("抓取到的所有数据:", len(result), result)
                return result
            except NoSuchElementException:
                element = driver.find_element_by_xpath(
                    '//*[@id="quality-dialog"]/div/div/stage-all-quality/div/div/ul/li/p')
                if element.text == "没有数据":
                    print("数据报告未刷新，等待刷新...")
                    time.sleep(2 * 64)  # 报告未刷新出来，睡眠2分钟
                    self.click_element('//*[@id="refreshQuality"]')  # 点击刷新按钮
                continue

    def crawl_one(self, pipeline_name):
        self.search_one(pipeline_name)
        result = self.get_report()

        # 数据规整化
        i, len_res = 1, len(result)
        check_dict = {}
        while i < len_res:
            if result[i] in CHECK_SET:
                print(result[i], "====", result[i + 1])
                check_dict[result[i]] = result[i + 1]
                i += 1
            i += 1

        self.switch_to_main()
        # 返回规整化的数据字典
        return check_dict

    def parse_one(self, pipeline_name, languages, project_name, name=None):
        # 一条流水线存在多个服务
        check_data, check_index = {}, {}
        language_list = languages.split(sep='、')
        if len(language_list) == 0:
            print("*" * 40, "该流水线未分服务组", "*" * 40)
            return

        for language in language_list:
            if language == "Go服务":
                check_data = GO_CHECK_DATA
                check_index = GO_EXCEL_INDEX
            elif language == "C&C++服务":
                check_data = C_CHECK_DATA
                check_index = C_EXCEL_INDEX
            elif language == "Python服务" and len(language_list) > 1:
                check_data = Python_CHECK_DATA
                check_index = Python_EXCEL_INDEX
            elif language == "Python服务" and len(language_list) == 1:
                check_data = Only_Python_CHECK_DATA
                check_index = Python_EXCEL_INDEX

            column = self.get_cols(sheet_name=language, project_name=project_name)
            self.handle_code_quality(pipeline_name, language, column, check_data, check_index)

            # Go服务的工程，加入 llt 待处理队列
            if language == "Go服务" and name:
                self.llt_projects.append((name, column))

    def handle_code_quality(self, pipeline_name, language, column, check_data, check_index):
        print("======== [数据抓取] =========")
        check_dict = self.crawl_one(pipeline_name)

        # print("======== [数据清洗与存储] =========")
        sheet_name = language
        self.save_hyperlink(sheet_name, column, 2, self.current_url)
        # 一个插件需要多个数据项
        for plugin, items in check_data.items():
            # 清洗出来每一个插件的需要的每一项数据
            for item_name in items:
                try:
                    check_dict[plugin]
                except KeyError:
                    # print(f"{plugin}——{item_name}", "KeyError")
                    continue
                else:
                    item_value = self.get_value(check_dict[plugin], item_name)
                    print(f"{plugin}——{item_name}", item_value)

                    # 存储到 EXCEL 表格
                    row = check_index[item_name]
                    if row != 0:
                        self.save_to_cell(sheet_name, column, row, float(item_value))

    def save_hyperlink(self, sheet_name, column, row, url):
        print(sheet_name, column, row, url)
        try:
            sheet = self.wb[sheet_name]
            try:
                sheet.cell(row=row, column=column).hyperlink = url
            except ValueError:
                print("存储失败，请检查数据项对应的表格行列是否正确")
        except KeyError:
            print("未在表格内找到对应的存储Sheet，请检查表格")

    def save_to_cell(self, sheet_name, column, row, value):
        print(sheet_name, column, row, value)
        try:
            sheet = self.wb[sheet_name]
            try:
                sheet.cell(row=row, column=column, value=value)
            except ValueError:
                print("存储失败，请检查数据项对应的表格行列是否正确")
        except KeyError:
            print("未在表格内找到对应的存储Sheet，请检查表格")

    def save_to_excel(self):
        try:
            self.wb.save(FINAL_OUTPUT_EXCEL_PATH)
        except PermissionError as error:
            print(error)
            print("======== [权限错误，存储失败，请检测是否关闭文件!!!] =========")
        except OSError as error:
            print(error)
            print("======== [存储失败，请检查文件格式!!!] =========")
        else:
            print("======== [所有数据存储至EXCEL表格已完成] =========")
            print(os.path.abspath(FINAL_OUTPUT_EXCEL_PATH))

    def get_cols(self, sheet_name, project_name):
        """
        获取工程名所在的列号
        :param sheet_name: EXCEL的 Sheet
        :param project_name: 工程名
        :return: 成功返回列号，否则返回-1
        """
        df = self.sheets_dict[sheet_name]
        cols = pd.DataFrame(data=np.arange(1, len(df.columns) + 1), index=df.iloc[0, :].values)
        try:
            col = cols.at[project_name, 0]
            return col
        except KeyError:
            return -1

    @staticmethod
    def get_value(string, key):
        """
        正则匹配获取数据项的值
        :param string: 爬取下来的包含数据项和数据项值的字符串
        :param key: 要匹配的数据项
        :return: 返回一个字符串类型的数字值
        """
        if not string:
            return 0
        value = re.findall(pattern=f"{key} (.*?) ", string=string)
        if value:
            return value[0]
        else:
            return re.findall(pattern=f"{key} (.*?)$", string=string)[0]

    def login_metabase(self):
        url = 'http://100.100.87.110:3000/question/14'
        self.open_url(url)

        print("Login Metabase...")
        self.send_keys('//*[@name="username"]', "coredata@huawei.com")
        self.send_keys('//*[@name="password"]', "APBKs6AJn5Qoga")
        self.click_element('//*[@id="root"]/div/div/div/div[2]/form/div[4]/button')
        time.sleep(5)
        cookies = self.browser.get_cookies()
        string = "=".join([cookies[0]["name"], cookies[0]["value"]])
        self.quit()
        return string

    def get_csv(self):
        # LLT看板登录账号密码：http://rnd-isource.huawei.com/c00466117/MyBlogs/tree/master/%E6%95%B0%E6%8D%AE%E7%BB%9F%E8%AE%A1/%E3%80%90%E6%9E%84%E5%BB%BA%E5%BA%A6%E9%87%8F%E3%80%91%E3%80%90%E5%B7%A5%E5%85%B7%E6%9C%88%E6%8A%A5%E3%80%91%E6%9E%84%E5%BB%BA%E8%AF%A6%E6%83%85%E5%8F%AF%E8%A7%86%E5%8C%96.md
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) "
                          "Chrome/84.0.4147.89 Safari/537.36",
            "Cookie": self.login_metabase(),
            "Origin": "http://100.100.87.110:3000",
            "Connection": "keep-alive",
            "Referer": "http://100.100.87.110:3000/question/14",

        }
        payload = {"parameters": ""}
        response = requests.post(url="http://100.100.87.110:3000/api/card/14/query/csv", data=payload, headers=headers)
        with open('tmp.csv', mode="w") as f:
            f.write(response.text.strip())

    def llt_df(self):
        self.get_csv()
        df = pd.read_csv("tmp.csv", usecols=["Date", "Module", "Version", "Llt Tc Changed Coverage",
                                             "Llt Tc Total Coverage", "Llt Tc Total Number",
                                             "Llt Problem Interception", "Llt Tc Pass Rate", "URL"])
        NFV_df = df[df["Version"].isin(["NFV_FusionStage 21.0.0"])]
        NFV_df = NFV_df.sort_values(by="Date", ascending=True).drop_duplicates(subset=["Module"], keep="last")
        NFV_df.index = NFV_df.Module
        NFV_df = NFV_df.drop(["Version", "Date", "Module"], axis=1)
        print(NFV_df)
        return NFV_df

    def handler_llt(self):
        print("将要处理的 llt 工程名和其所在的 EXCEL 列号")
        print(self.llt_projects)
        if not self.llt_projects:
            return

        df = self.llt_df()
        name_set = set(list(df.index))
        for x in self.llt_projects:
            job_name, column = x[0], x[1]
            if job_name not in name_set:
                continue
            print(job_name)
            Llt_Tc_Changed_Coverage = float(df.at[job_name, "Llt Tc Changed Coverage"]) * 100
            Llt_Tc_Total_Coverage = float(df.at[job_name, "Llt Tc Total Coverage"]) * 100
            Llt_Tc_Total_Number = df.at[job_name, "Llt Tc Total Number"]
            Llt_Problem_Interception = df.at[job_name, "Llt Problem Interception"]
            Llt_Tc_Pass_Rate = float(df.at[job_name, "Llt Tc Pass Rate"]) * 100
            Llt_URL = df.at[job_name, "URL"]

            self.save_to_cell("Go服务", column, GO_EXCEL_INDEX["Llt Tc Changed Coverage"], Llt_Tc_Changed_Coverage)
            self.save_to_cell("Go服务", column, GO_EXCEL_INDEX["Llt Tc Total Coverage"], Llt_Tc_Total_Coverage)
            self.save_to_cell("Go服务", column, GO_EXCEL_INDEX["Llt Tc Total Number"], Llt_Tc_Total_Number)
            self.save_to_cell("Go服务", column, GO_EXCEL_INDEX["Llt Problem Interception"], Llt_Problem_Interception)
            self.save_to_cell("Go服务", column, GO_EXCEL_INDEX["Llt Tc Pass Rate"], Llt_Tc_Pass_Rate)
            # self.save_to_cell("Go服务", column, GO_EXCEL_INDEX["Llt URL"], job_name)
            self.save_hyperlink("Go服务", column, GO_EXCEL_INDEX["Llt Tc Pass Rate"], Llt_URL)


def run():
    """
    运行函数
    :return:
    """
    # 初始化
    win = CloudPiplineWebUI()
    url = 'https://szxy1-cdt.huawei.com/pipeline/home'
    win.open_url(url)
    win.login()
    win.select_all_pipelines()

    # 1. 取得流水线的所有信息
    # 读取excel文件，并指定Sheet
    pipeline_df = pd.read_excel(f"{PIPELINES_OUTPUT_EXCEL_PATH}.xlsx", header=0, index_col=0, )
    for index, row in pipeline_df.iterrows():
        pipeline_name, languages, project_name, team, xm, name = row["质量流水线"], row["语言"], row["工程名称"], row["维护团队"], row[
            "归属 XM团队"], row["Sonar名称"]

        # # 精确定位某条流水线示例：
        # if project_name != "ALARMNOTIFYSERVICE" and project_name != "CFETOOLSBOX":
        #     continue

        print(f"第 【{index + 1}】 条流水线 [ {pipeline_name} ] 正在抓取中...")
        # 2. 爬取数据项
        win.parse_one(pipeline_name=pipeline_name, languages=languages, project_name=project_name, name=name)

    win.handler_llt()

    win.quit()

    # 保存值到excel
    win.save_to_excel()


if __name__ == '__main__':
    run()
