
from Config.settings import GO_CHECK_DATA_CN, Go_EXCEL_INDEX, C_CHECK_DATA_CN, C_EXCEL_INDEX
from Config.settings import Python_CHECK_DATA_CN, Python_EXCEL_INDEX, Only_Python_CHECK_DATA_CN
from Config.settings import FINAL_OUTPUT_EXCEL_PATH, PIPELINES_OUTPUT_EXCEL_PATH, EXECUTABLE_PATH, INPUT_EXCEL_PATH
from CodeCD.corecd_driver import CoreCD_Driver
import time
import re
import json
from itertools import dropwhile, islice
import pandas as pd
import numpy as np
import openpyxl
import os
from Log.logging import Logger
logger = Logger()


class CoreSpider(CoreCD_Driver):
    def __init__(self, show_windows=True):
        super().__init__(show_windows=show_windows)
        self.sheets_dict = {}
        self.current_url = ''
        self.wb = openpyxl.load_workbook(INPUT_EXCEL_PATH)
        self.init_integration_excel_df()

    def init_integration_excel_df(self):
        sheets = ["Go服务", "Python服务", "C&C++服务"]
        for s in sheets:
            info_integration_df = pd.read_excel(INPUT_EXCEL_PATH, sheet_name=s)
            self.sheets_dict[s] = info_integration_df

    def parse_one(self, pipeline_name, languages, project_name, check_dict=None):
        # 一条流水线存在多个服务
        language_list = languages.split(sep='、')
        if len(language_list) == 0:
            print("*" * 40, "该流水线未分服务组", "*" * 40)
            return

        check_data, check_index = {}, {}
        for language in language_list:
            if language == "Go服务":
                check_data = GO_CHECK_DATA_CN
                check_index = Go_EXCEL_INDEX
            elif language == "C&C++服务":
                check_data = C_CHECK_DATA_CN
                check_index = C_EXCEL_INDEX
            elif language == "Python服务" and len(language_list) > 1:
                check_data = Python_CHECK_DATA_CN
                check_index = Python_EXCEL_INDEX
            elif language == "Python服务" and len(language_list) == 1:
                check_data = Only_Python_CHECK_DATA_CN
                check_index = Python_EXCEL_INDEX

            sheet_name = language
            column = self.get_cols(sheet_name=language, project_name=project_name)
            print("所在Sheet:", sheet_name, " 所在列:", column)
            # 一个插件需要多个数据项
            for plugin, items in check_data.items():
                # 清洗出来每一个插件需要的每一项数据
                for item_name in items:
                    try:
                        check_dict[plugin]
                    except KeyError:
                        # print(f"{plugin}——{item_name}", "KeyError")
                        continue
                    else:
                        item_value = check_dict[plugin][item_name]
                        print(f"{plugin}——{item_name}", item_value)

                        # 存储到 EXCEL 表格
                        row = check_index[item_name]
                        if row != 0:
                            self.save_to_cell(sheet_name, column, row, float(item_value))
            self.save_to_cell(sheet_name, column, 22, self.current_url)       # TODO 此处存在了22列

    def save_hyperlink(self, sheet_name, column, row, url):
        print(sheet_name, column, row, url)
        try:
            sheet = self.wb[sheet_name]
            try:
                sheet.cell(row=row, column=column).hyperlink = url
            except ValueError:
                print("存储失败，请检查数据项对应的表格行列是否正确")
        except KeyError:
            print("未在表格内找到对应的存储Sheet，请检查表格")

    def save_to_cell(self, sheet_name, column, row, value):
        print(sheet_name, column, row, value)
        try:
            sheet = self.wb[sheet_name]
            try:
                sheet.cell(row=row, column=column, value=value)
            except ValueError:
                print("存储失败，请检查数据项对应的表格行列是否正确")
        except KeyError:
            print("未在表格内找到对应的存储Sheet，请检查表格")

    def save_to_excel(self):
        try:
            self.wb.save(FINAL_OUTPUT_EXCEL_PATH)
        except PermissionError as error:
            print(error)
            print("======== [权限错误，存储失败，请检测是否关闭文件!!!] =========")
        except OSError as error:
            print(error)
            print("======== [存储失败，请检查文件格式!!!] =========")
        else:
            print("======== [所有数据存储至EXCEL表格已完成] =========")
            print(os.path.abspath(FINAL_OUTPUT_EXCEL_PATH))

    def get_cols(self, sheet_name, project_name):
        """
        获取工程名所在的列号
        :param sheet_name: EXCEL的 Sheet
        :param project_name: 工程名
        :return: 成功返回列号，否则返回-1
        """
        df = self.sheets_dict[sheet_name]
        cols = pd.DataFrame(data=np.arange(1, len(df.columns) + 1), index=df.iloc[0, :].values)
        try:
            col = cols.at[project_name, 0]
            return col
        except KeyError:
            return -1

    def get_last_record(self, pipeline_name):
        template_id = self.getter.get_template_id(pipeline_name)
        self.current_url = self.get_template_run_detail_url(template_id)
        data = self.getter.get_last_record(template_id)
        return data

    @staticmethod
    def qingxi(data):
        recordSubChartInfos = data['data']['recordSubChartInfos']
        check_data = {}
        for record in recordSubChartInfos:
            chartNodes = record["chartNode"]
            for chartNode in islice(chartNodes, 2, None):
                build_info = chartNode["build_info"]
                if build_info is None:
                    continue
                runstate = chartNode["runstate"]
                log_url = chartNode["log_url"]
                job_type = chartNode["job_type"]
                label = chartNode["label"]
                build_info = json.loads(build_info)
                check_one = {}
                for i in build_info:
                    check_one[i['cn_name']] = i['actual']
                check_data[label] = check_one
        for k, v in check_data.items():
            print(k)
            for k1, v1 in v.items():
                print('\t', k1, v1)

        return check_data


def run():
    my_driver = CoreSpider(show_windows=False)
    my_driver.login()
    my_driver.open_pipeline_page()

    # 1. 取得流水线的所有信息
    # 读取excel文件，并指定Sheet
    pipeline_df = pd.read_excel(PIPELINES_OUTPUT_EXCEL_PATH, header=0,)
    for index, row in pipeline_df.iterrows():
        pipeline_name, languages, project_name, xm = row["质量流水线"], row["语言"], row["工程名称"], row["归属 XM团队"]

        print(f"第 【{index}】 条流水线 [ {pipeline_name} ] 正在抓取中...")

        # 2. 爬取数据项
        data = my_driver.get_last_record(pipeline_name)

        # 3. 清洗数据
        item = my_driver.qingxi(data)

        # 4. 存储数据
        my_driver.parse_one(pipeline_name=pipeline_name, languages=languages, project_name=project_name, check_dict=item)

    # 保存值到excel
    my_driver.save_to_excel()


if __name__ == '__main__':
    run()

